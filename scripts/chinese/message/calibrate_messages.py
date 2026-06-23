#!/usr/bin/env python3
"""
Message calibration tool for SoH Chinese translation.

Reads:
  - message_raw_ntsc.txt          (English NES message raw hex data)
  - message_raw_cn.txt             (Chinese iQue message raw hex data)
  - OOT_SimplifiedChinese_TextDump.html  (N64 iQue Simplified Chinese text dump)
  - charmap_ntsc.txt               (NTSC charmap: extended Latin + button icons)
  - charmap_chn.txt                (Chinese charmap: iQue codes + button icons)

Generates an Excel workbook with:
  Col A: textId
  Col B: NTSC raw hex
  Col C: NTSC readable (control codes preserved as [0x##])
  Col D: CN raw hex
  Col E: CN readable
  Col F: OOT dump text

Comparisons (highlighted):
  - Control code mismatch (NTSC vs CN): bright red background
  - Chinese character mismatch (CN vs OOT dump): yellow background

Usage:
  uv run message/calibrate_messages.py
"""

import re
import struct
import html as html_mod
from pathlib import Path
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. Control character definitions
# ---------------------------------------------------------------------------

# Control code name lookup (from message_data_fmt.h)
CTRL_NAMES = {
    0x01: "NEWLINE",
    0x02: "END",
    0x04: "BOX_BREAK",
    0x05: "COLOR",
    0x06: "SHIFT",
    0x07: "TEXTID",
    0x08: "QUICKTEXT_ENABLE",
    0x09: "QUICKTEXT_DISABLE",
    0x0A: "PERSISTENT",
    0x0B: "EVENT",
    0x0C: "BOX_BREAK_DELAYED",
    0x0D: "AWAIT_BUTTON_PRESS",
    0x0E: "FADE",
    0x0F: "NAME",
    0x10: "OCARINA",
    0x11: "FADE2",
    0x12: "SFX",
    0x13: "ITEM_ICON",
    0x14: "TEXT_SPEED",
    0x15: "BACKGROUND",
    0x16: "MARATHON_TIME",
    0x17: "RACE_TIME",
    0x18: "POINTS",
    0x19: "TOKENS",
    0x1A: "UNSKIPPABLE",
    0x1B: "TWO_CHOICE",
    0x1C: "THREE_CHOICE",
    0x1D: "FISH_INFO",
    0x1E: "HIGHSCORE",
    0x1F: "TIME",
}

# Number of parameter bytes following each control code
CTRL_PARAMS = {
    0x01: 0, 0x02: 0, 0x04: 0, 0x05: 1, 0x06: 1,
    0x07: 2, 0x08: 0, 0x09: 0, 0x0A: 0, 0x0B: 0,
    0x0C: 1, 0x0D: 0, 0x0E: 1, 0x0F: 0, 0x10: 0,
    0x11: 2, 0x12: 2, 0x13: 1, 0x14: 1, 0x15: 3,
    0x16: 0, 0x17: 0, 0x18: 0, 0x19: 0, 0x1A: 0,
    0x1B: 0, 0x1C: 0, 0x1D: 0, 0x1E: 1, 0x1F: 0,
}

# Set of 1-byte control codes (codes <= 0x1F plus button range 0x9F-0xAB for NTSC)
CTRL_SINGLE_BYTE_SET = set(CTRL_PARAMS.keys())

# ---------------------------------------------------------------------------
# 2. Parse charmaps (hex code → display char)
# ---------------------------------------------------------------------------

def parse_charmap_ntsc(path: str) -> dict[int, str]:
    """Parse charmap_ntsc.txt: 'DisplayChar: 0xCode' or '(A): 0x9F'

    Returns: {0x80: 'À', ..., 0x9F: '(A)', ...}
    """
    charmap = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            m = re.match(r"^'?(.+?)'?\s*:\s*0x([0-9A-Fa-f]+)", line)
            if m:
                char_str = m.group(1)
                code = int(m.group(2), 16)
                charmap[code] = char_str
    return charmap


def parse_charmap_chn(path: str) -> dict[int, str]:
    """Parse charmap_chn.txt: simple format 'CHAR: 0xCode' (no quotes).

    Examples:
        (A): 0xAA9F
        你: 0xA08C
        (↑): 0xAAA5    (arrow buttons)

    Returns: {0xA08C: '你', ..., 0xAA9F: '(A)', ...}
    All button icons are already in parentheses format (A), (↑), etc.
    """
    charmap = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # Match: <character_or_button_representation>: 0x<hex>
            # The representation is everything before the colon
            m = re.match(r"^(.+?)\s*:\s*0x([0-9A-Fa-f]+)", line)
            if m:
                char_str = m.group(1).strip()
                code = int(m.group(2), 16)
                charmap[code] = char_str
    return charmap


# ---------------------------------------------------------------------------
# 3. Parse message raw files
# ---------------------------------------------------------------------------

def parse_message_raw(path: str) -> dict[int, list[int]]:
    """Parse message_raw_*.txt: 0x0001 = { 0x01, 0x02, ... };

    Returns: {textId: [byte, byte, ...]}
    """
    messages = {}
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    for m in re.finditer(r"0x([0-9A-Fa-f]{4})\s*=\s*\{\s*([^}]*?)\s*\};", content):
        text_id = int(m.group(1), 16)
        hex_bytes_str = m.group(2)
        byte_values = []
        for bm in re.finditer(r"0x([0-9A-Fa-f]{2})", hex_bytes_str):
            byte_values.append(int(bm.group(1), 16))
        messages[text_id] = byte_values
    return messages


# ---------------------------------------------------------------------------
# 4. Parse OOT dump HTML
# ---------------------------------------------------------------------------

def parse_oot_dump_html(path: str) -> dict[int, str]:
    """Parse OOT_SimplifiedChinese_TextDump.html.

    Extracts textId → plain text content (strips HTML tags, preserves
    button/control representations like (A), (t), etc.)
    """
    import html as html_mod

    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    messages = {}
    # Find the main table body — text entries are in <tr><td> blocks starting with 0x...
    # Pattern: <td id="0002">0x0002:<br>...content...</td> or <td>0x0001:<br>...content...</td>

    # Split on <tr> to get rows, then find <td> with textId
    rows = re.split(r'<tr[^>]*>', content)

    for row in rows:
        # Extract textId from id attribute or from 0xXXXX: prefix
        td_match = re.search(r'<td[^>]*?id\s*=\s*"([0-9A-Fa-f]{4})"', row)
        if td_match:
            text_id = int(td_match.group(1), 16)
        else:
            # Try matching 0xXXXX: pattern
            td_match = re.search(r'<td[^>]*?>?\s*0x([0-9A-Fa-f]{4})\s*:', row)
            if td_match:
                text_id = int(td_match.group(1), 16)
            else:
                continue

        # Extract text content: everything between the first <br> or after 0xXXXX:<br> and </td>
        # Remove all HTML tags, keep text content
        # First, get everything inside the <td>
        td_content_match = re.search(r'<td[^>]*?>(.*?)</td>', row, re.DOTALL)
        if not td_content_match:
            continue

        td_content = td_content_match.group(1)

        # Remove the "0xXXXX:<br>" header
        td_content = re.sub(r'^\s*0x[0-9A-Fa-f]{4}\s*:\s*<br[^>]*?>', '', td_content)
        # Remove the "0xXXXX:<br>" variant without <br>
        td_content = re.sub(r'^\s*0x[0-9A-Fa-f]{4}\s*:\s*', '', td_content)

        # Strip all HTML tags but preserve text
        # Replace <br> with newline
        td_content = re.sub(r'<br[^>]*?>', '\n', td_content)
        # Remove all other tags
        td_content = re.sub(r'<[^>]*?>', '', td_content)
        # Decode HTML entities
        td_content = html_mod.unescape(td_content)
        # Normalize whitespace
        lines = [l.strip() for l in td_content.split('\n')]
        td_content = '\n'.join(line for line in lines if line)

        # Avoid duplicate keys (some messages appear multiple times, keep last)
        messages[text_id] = td_content

    return messages


# ---------------------------------------------------------------------------
# 5. Convert raw bytes → control code string (for comparison)
# ---------------------------------------------------------------------------

# Control codes to IGNORE when comparing NTSC vs CN control sequences.
# These are formatting-only codes that naturally differ between languages
# due to CJK vs ASCII character density and layout preferences.
IGNORED_CTRL_CODES = {0x01, 0x04, 0x05, 0x06, 0x0C, 0x13, 0x1A}  # NEWLINE, BOX_BREAK, COLOR, SHIFT, BOX_BREAK_DELAYED, ITEM_ICON, UNSKIPPABLE


def extract_control_sequence(data: list[int], ignore: set[int] = None,
                             is_cn: bool = False) -> str:
    """Extract only the control codes (0x01-0x1F) and their parameters from raw data.

    Returns a normalized string like: "1A,13:2D,08,05:41,05:40,09,02"
    where each control code is followed by its parameters after ':'.

    If `ignore` is set, those control codes (and their parameters) are skipped.
    Defaults to IGNORED_CTRL_CODES for CN-vs-NTSC comparison.

    If `is_cn` is True, bytes >= 0xA0 are treated as 2-byte CJK sequence
    prefixes and skipped along with their low byte.  For NTSC data this
    must be False because 0x9F-0xAB are single-byte button icons.
    """
    if ignore is None:
        ignore = IGNORED_CTRL_CODES
    parts = []
    i = 0
    while i < len(data):
        b = data[i]
        # Only for CN: >= 0xA0 are high bytes of 2-byte CJK sequences
        if is_cn and (b >= 0xA0) and i + 1 < len(data):
            i += 2
            continue
        if b in CTRL_PARAMS:
            nparams = CTRL_PARAMS[b]
            params = data[i+1:i+1+nparams]
            if b not in ignore:
                param_str = ":".join(f"{p:02X}" for p in params)
                if param_str:
                    parts.append(f"{b:02X}:{param_str}")
                else:
                    parts.append(f"{b:02X}")
            i += 1 + nparams
        else:
            i += 1
    return ",".join(parts)


# ---------------------------------------------------------------------------
# 6. Convert raw bytes → readable text
# ---------------------------------------------------------------------------

def raw_to_readable_ntsc(data: list[int], charmap: dict[int, str]) -> str:
    """Convert NTSC raw bytes to readable string.

    - Control codes: [0x## param param]
    - ASCII 0x20-0x7E: display as the character
    - Extended Latin 0x80-0x9E: use charmap
    - Button icons 0x9F-0xAB: use charmap
    """
    result = []
    i = 0
    while i < len(data):
        b = data[i]
        if b in CTRL_PARAMS:
            nparams = CTRL_PARAMS[b]
            params = data[i+1:i+1+nparams]
            param_str = " ".join(f"0x{p:02X}" for p in params)
            name = CTRL_NAMES.get(b, "??")
            if param_str:
                result.append(f"[0x{b:02X} {param_str}]")
            else:
                result.append(f"[0x{b:02X}]")
            i += 1 + nparams
        elif b in charmap:
            result.append(charmap[b])
            i += 1
        elif 0x20 <= b <= 0x7E:
            result.append(chr(b))
            i += 1
        elif b == 0x7F:
            result.append(" ")  # blank
            i += 1
        else:
            # Unknown byte — show as hex
            result.append(f"\\x{b:02X}")
            i += 1
    return "".join(result)


def raw_to_readable_cn(data: list[int], charmap: dict[int, str]) -> str:
    """Convert Chinese raw bytes to readable string.

    - Control codes: [0x## param param]
    - >= 0xA0 + lowByte: 2-byte Chinese character/button (use charmap)
    - ASCII 0x20-0x7E: display as the character
    """
    result = []
    i = 0
    while i < len(data):
        b = data[i]

        # 2-byte Chinese character (high byte >= 0xA0)
        # Valid ranges: 0xA08C-0xA775 (main iQue), 0xAA9F-0xAAAB (buttons),
        #               0xAAAC-0xAC48 (extended v8)
        if (b >= 0xA0) and i + 1 < len(data):
            low = data[i + 1]
            chi_code = (b << 8) | low
            if chi_code in charmap:
                result.append(charmap[chi_code])
            elif 0xAA9F <= chi_code <= 0xAAAB:
                # Button icon not in charmap — use hex fallback
                result.append(f"\\x{chi_code:04X}")
            else:
                result.append(f"\\x{chi_code:04X}")
            i += 2
            continue

        # Control codes
        if b in CTRL_PARAMS:
            nparams = CTRL_PARAMS[b]
            params = data[i+1:i+1+nparams]
            param_str = " ".join(f"0x{p:02X}" for p in params)
            if param_str:
                result.append(f"[0x{b:02X} {param_str}]")
            else:
                result.append(f"[0x{b:02X}]")
            i += 1 + nparams
        elif 0x20 <= b <= 0x7E:
            result.append(chr(b))
            i += 1
        elif b in charmap:
            result.append(charmap[b])
            i += 1
        else:
            result.append(f"\\x{b:02X}")
            i += 1
    return "".join(result)


# ---------------------------------------------------------------------------
# 7. Chinese character extraction with newline preservation (for comparison)
# ---------------------------------------------------------------------------

def extract_cn_chars_with_breaks(data: list[int], charmap: dict[int, str]) -> str:
    """Extract Chinese characters from CN raw bytes, preserving line/box breaks.

    - 0x01 (NEWLINE)          → '\n'
    - 0x04 (BOX_BREAK)        → '\n'
    - 0x0C (BOX_BREAK_DELAYED) → '\n' (delay param is consumed)
    - >= 0xA0 + lowByte     → Chinese char or button icon (from charmap)
    - All other control codes  → skipped (with their params)
    - ASCII 0x20-0x7E          → skipped
    """
    result = []
    i = 0
    while i < len(data):
        b = data[i]

        # 2-byte Chinese character / button icon
        if (b >= 0xA0) and i + 1 < len(data):
            low = data[i + 1]
            chi_code = (b << 8) | low
            if chi_code in charmap:
                result.append(charmap[chi_code])
            i += 2
            continue

        # Line/box breaks → '\n'
        if b in (0x01, 0x04, 0x0C):
            result.append('\n')
            nparams = CTRL_PARAMS.get(b, 0)
            i += 1 + nparams
            continue

        # Other control codes → skip with params
        if b in CTRL_PARAMS:
            nparams = CTRL_PARAMS[b]
            i += 1 + nparams
            continue

        # ASCII digits and space → keep (meaningful in Chinese messages: "20枚", "80卢比")
        if (0x30 <= b <= 0x39) or b == 0x20:
            result.append(chr(b))
            i += 1
            continue

        # Anything else (ASCII letters, extended Latin, etc.) → skip
        i += 1

    result_str = "".join(result)
    # Collapse consecutive newlines into one — only care about WHERE, not how many
    result_str = re.sub(r'\n+', '\n', result_str)
    # Trim spaces around newlines — iQue sometimes has trailing/leading spaces
    result_str = re.sub(r' *\n *', '\n', result_str)
    # Strip leading/trailing spaces (e.g. space before 0x02 END in iQue raw)
    result_str = result_str.strip()
    return result_str


def extract_oot_chars_with_breaks(oot_text: str) -> str:
    """Extract Chinese characters from OOT dump text, preserving newlines.

    OOT dump already has '\n' from <br> tags. We keep:
    - Chinese characters (CJK Unified, Extension A, CJK punctuation, fullwidth)
    - Button representations: (A), (B), (C), (L), (R), (Z),
      (↑), (↓), (←), (→), (▼), (+), (P)
    - Newlines '\n'
    - Other parenthesized items are skipped (like (t), (m), (r), (s),
      (Highscore: ...), (只是现价), etc.)
    Before extraction, strip control-code placeholders the dump uses for
    runtime values (not present in iQue raw data):
      (t)=TIME (n)=NAME (m)=MARATHON (r)=RACE (p)=POINTS (s)=TOKENS
      (Highscore: ...) and other parenthetical notes ≥3 chars.
    """
    # Pre-clean runtime placeholders — these don't exist in iQue raw
    cleaned = oot_text
    cleaned = re.sub(r'\([tmnrps]\)', '', cleaned)               # single-letter
    cleaned = re.sub(r'\(Highscore:[^)]*\)', '', cleaned)        # highscore

    result = []
    i = 0
    while i < len(cleaned):
        ch = cleaned[i]
        cp = ord(ch)

        # Chinese character (including CJK-compatible punctuation)
        if (0x4E00 <= cp <= 0x9FFF or       # CJK Unified
            0x3400 <= cp <= 0x4DBF or       # CJK Extension A
            0x3000 <= cp <= 0x303F or       # CJK Punctuation
            0xFF00 <= cp <= 0xFFEF or       # Fullwidth forms
            0xF900 <= cp <= 0xFAFF or       # CJK Compatibility
            0x2000 <= cp <= 0x206F or       # General Punctuation (… — etc.)
            cp == 0x30FB):                   # Katakana middle dot (・) used in CN names
            result.append(ch)
            i += 1
            continue

        # Newline
        if ch == '\n':
            result.append('\n')
            i += 1
            continue

        # Digit or space → keep (meaningful in Chinese messages)
        if ch.isdigit() or ch == ' ':
            result.append(ch)
            i += 1
            continue

        # Button icon: (A), (B), (C), (L), (R), (Z), (↑), (↓), (←), (→), (▼), (+), (P)
        if ch == '(' and i + 2 < len(cleaned):
            close = cleaned[i + 2]
            inner = cleaned[i + 1]
            if close == ')':
                token = cleaned[i:i+3]
                btn_single_set = set('ABCLRZ↑↓←→▼+P')
                if inner in btn_single_set:
                    result.append(token)
                    i += 3
                    continue
                # Skip non-button parenthesized items like (t), (m), (r), (s), (n), (p)
                i += 3
                continue

        i += 1

    # Normalize: collapse consecutive newlines into one, strip leading/trailing,
    # and trim spaces around newlines (iQue vs OOT spacing inconsistency)
    result_str = "".join(result)
    result_str = re.sub(r'\n+', '\n', result_str)
    result_str = re.sub(r' *\n *', '\n', result_str)
    result_str = result_str.strip()
    return result_str


# ---------------------------------------------------------------------------
# 8. Main build Excel function
# ---------------------------------------------------------------------------

FILL_RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")       # control mismatch
FILL_YELLOW = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")     # char mismatch
FILL_LIGHT_GRAY = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_excel(
    ntsc_messages: dict[int, list[int]],
    cn_messages: dict[int, list[int]],
    oot_dump: dict[int, str],
    charmap_ntsc: dict[int, str],
    charmap_chn: dict[int, str],
    output_path: str,
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Message Calibration"

    # Header row
    headers = [
        "textId",
        "NTSC Raw Hex",
        "NTSC Readable",
        "CN Raw Hex",
        "CN Readable",
        "N64 CN Text Dump",
        "Ctrl Match",
        "Char Match",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Collect all textIds from all sources
    all_ids = sorted(set(ntsc_messages.keys()) | set(cn_messages.keys()) | set(oot_dump.keys()))

    row = 2
    ctrl_mismatch_count = 0
    char_mismatch_count = 0

    for text_id in all_ids:
        ntsc_data = ntsc_messages.get(text_id)
        cn_data = cn_messages.get(text_id)
        oot_text = oot_dump.get(text_id, "")

        # --- Raw hex strings ---
        ntsc_raw = ", ".join(f"0x{b:02X}" for b in ntsc_data) if ntsc_data else "(missing)"
        cn_raw = ", ".join(f"0x{b:02X}" for b in cn_data) if cn_data else "(missing)"

        # --- Readable strings ---
        ntsc_readable = raw_to_readable_ntsc(ntsc_data, charmap_ntsc) if ntsc_data else "(missing)"
        cn_readable = raw_to_readable_cn(cn_data, charmap_chn) if cn_data else "(missing)"

        # --- Comparison 1: Control codes (ignoring formatting-only codes) ---
        # Normalize: NTSC messages sometimes lack trailing 0x02 (end with TEXTID
        # instead); iQue always appends it. Ensure both sides have it for fair comparison.
        if ntsc_data and ntsc_data[-1] != 0x02:
            ntsc_data = list(ntsc_data) + [0x02]
        if cn_data and cn_data[-1] != 0x02:
            cn_data = list(cn_data) + [0x02]

        ntsc_ctrl = extract_control_sequence(ntsc_data, is_cn=False) if ntsc_data else ""
        cn_ctrl = extract_control_sequence(cn_data, is_cn=True) if cn_data else ""
        ctrl_match = (ntsc_ctrl == cn_ctrl)

        # --- Comparison 2: Chinese characters with line/box breaks ---
        cn_chars_with_nl = extract_cn_chars_with_breaks(cn_data, charmap_chn) if cn_data else ""
        oot_chars_with_nl = extract_oot_chars_with_breaks(oot_text) if oot_text else ""
        char_match = (cn_chars_with_nl == oot_chars_with_nl)

        # Fuzzy match: iQue and OOT dump may differ only in word-choice
        # preferences.  Normalize both sides to the same variant.
        if not char_match:
            cn_norm = cn_chars_with_nl
            oot_norm = oot_chars_with_nl
            for a, b in [('或', '和'), ('他', '她'), ('象', '像'),('盔甲', '铠甲')]:
                cn_norm = cn_norm.replace(a, b)
                oot_norm = oot_norm.replace(a, b)
            char_match = (cn_norm == oot_norm)

        # Write data
        ws.cell(row=row, column=1, value=f"0x{text_id:04X}").border = THIN_BORDER
        ws.cell(row=row, column=2, value=ntsc_raw).border = THIN_BORDER
        ws.cell(row=row, column=3, value=ntsc_readable).border = THIN_BORDER
        ws.cell(row=row, column=4, value=cn_raw).border = THIN_BORDER
        ws.cell(row=row, column=5, value=cn_readable).border = THIN_BORDER
        ws.cell(row=row, column=6, value=oot_text).border = THIN_BORDER

        ctrl_cell = ws.cell(row=row, column=7, value="✓" if ctrl_match else "✗ MISMATCH")
        ctrl_cell.border = THIN_BORDER
        char_cell = ws.cell(row=row, column=8, value="✓" if char_match else "✗ MISMATCH")
        char_cell.border = THIN_BORDER

        # --- Highlight mismatches ---
        if not ctrl_match:
            ctrl_mismatch_count += 1
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = FILL_RED
        elif not char_match:
            char_mismatch_count += 1
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = FILL_YELLOW
        elif text_id % 2 == 1:
            # Alternate row shading for matched entries
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = FILL_LIGHT_GRAY

        row += 1

    # Column widths
    col_widths = [10, 50, 60, 50, 60, 55, 12, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:H{row-1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Total Messages").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=len(all_ids))
    ws2.cell(row=2, column=1, value="Control Code Mismatches").font = Font(bold=True)
    ws2.cell(row=2, column=2, value=ctrl_mismatch_count)
    ws2.cell(row=2, column=2).font = Font(color="FF0000", bold=True) if ctrl_mismatch_count else Font()
    ws2.cell(row=3, column=1, value="Chinese Character Mismatches").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=char_mismatch_count)
    ws2.cell(row=3, column=2).font = Font(color="FF8C00", bold=True) if char_mismatch_count else Font()
    ws2.cell(row=5, column=1, value="Legend:").font = Font(bold=True)
    ws2.cell(row=6, column=1, value="RED background").fill = FILL_RED
    ws2.cell(row=6, column=2, value="= Control code mismatch (NTSC vs CN)")
    ws2.cell(row=7, column=1, value="YELLOW background").fill = FILL_YELLOW
    ws2.cell(row=7, column=2, value="= Chinese character mismatch (CN vs OOT dump)")
    ws2.cell(row=8, column=1, value="GRAY background").fill = FILL_LIGHT_GRAY
    ws2.cell(row=8, column=2, value="= Matched (alternating rows)")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 55

    wb.save(output_path)
    print(f"Wrote {output_path}")
    print(f"Total: {len(all_ids)} messages")
    print(f"Control code mismatches (RED): {ctrl_mismatch_count}")
    print(f"Chinese character mismatches (YELLOW): {char_mismatch_count}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    here = Path(__file__).resolve().parent

    print("Parsing charmaps...")
    charmap_ntsc = parse_charmap_ntsc(str(here / "charmap" / "charmap_ntsc.txt"))
    print(f"  NTSC charmap: {len(charmap_ntsc)} entries")
    charmap_chn = parse_charmap_chn(str(here / "charmap" / "charmap_chn.txt"))
    print(f"  CN  charmap: {len(charmap_chn)} entries")

    print("Parsing message raw files...")
    ntsc_msgs = parse_message_raw(str(here / "txt" / "message_raw_ntsc.txt"))
    print(f"  NTSC messages: {len(ntsc_msgs)}")
    cn_msgs = parse_message_raw(str(here / "txt" / "message_raw_cn_from_ique.txt"))
    print(f"  CN   messages: {len(cn_msgs)}")

    print("Parsing OOT dump HTML...")
    oot_dump = parse_oot_dump_html(str(here / "calibration" / "OOT_SimplifiedChinese_TextDump.html"))
    print(f"  OOT dump entries: {len(oot_dump)}")

    output = here / "calibration" / "message_calibration.xlsx"
    build_excel(ntsc_msgs, cn_msgs, oot_dump, charmap_ntsc, charmap_chn, str(output))


if __name__ == "__main__":
    main()
